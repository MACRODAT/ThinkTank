"""api/routes/files.py — File drop, mining, and attachment endpoint."""
from __future__ import annotations
import uuid, json, base64, logging
from pathlib import Path
from datetime import datetime
from fastapi import APIRouter, Body, Request
import aiosqlite
from core.database import DB_PATH

router  = APIRouter(tags=["files"])
logger  = logging.getLogger(__name__)

UPLOADS = Path(__file__).parent.parent.parent / "data" / "uploads"
UPLOADS.mkdir(parents=True, exist_ok=True)


@router.post("/api/files/drop")
async def drop_file(request: Request):
    """
    Accept a base64-encoded file. Mine content and store summary.
    Supports: .txt, .md, .csv, .json, .xlsx (via openpyxl if available).
    """
    data     = await request.json()
    filename = data.get("filename", "file.txt")
    b64      = data.get("content_b64", "")
    mime     = data.get("mime", "text/plain")

    try:
        raw_bytes = base64.b64decode(b64)
    except Exception as e:
        return {"error": f"Base64 decode failed: {e}"}

    # Write to disk
    fid   = str(uuid.uuid4())
    fpath = UPLOADS / f"{fid}_{filename}"
    fpath.write_bytes(raw_bytes)

    # Mine content
    content  = ""
    summary  = ""
    metadata = {}
    ext      = Path(filename).suffix.lower()

    try:
        if ext in (".txt", ".md"):
            content = raw_bytes.decode("utf-8", errors="replace")
            lines   = content.splitlines()
            summary = f"{len(lines)} lines, {len(content)} chars"
            metadata = {"lines": len(lines), "chars": len(content)}

        elif ext == ".json":
            parsed  = json.loads(raw_bytes.decode("utf-8", errors="replace"))
            content = json.dumps(parsed, indent=2)[:5000]
            summary = f"JSON with {len(parsed) if isinstance(parsed, list) else len(parsed.keys()) if isinstance(parsed, dict) else 1} items"
            metadata = {"type": type(parsed).__name__}

        elif ext == ".csv":
            content = raw_bytes.decode("utf-8", errors="replace")
            lines   = content.splitlines()
            summary = f"CSV with {len(lines)} rows"
            if lines:
                headers = lines[0].split(",")
                metadata = {"rows": len(lines)-1, "columns": headers}

        elif ext in (".xlsx", ".xls"):
            try:
                import io
                import openpyxl
                wb    = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
                rows_data = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(max_row=100, values_only=True):
                        rows_data.append(list(row))
                content  = json.dumps(rows_data, default=str)[:8000]
                summary  = f"Excel: {len(wb.sheetnames)} sheets"
                metadata = {"sheets": wb.sheetnames, "rows_sampled": len(rows_data)}
            except ImportError:
                summary = "Excel file — install openpyxl to mine data"
                content = ""
            except Exception as ex:
                summary = f"Excel parse error: {ex}"

        elif ext == ".pdf":
            try:
                import io
                import PyPDF2
                reader  = PyPDF2.PdfReader(io.BytesIO(raw_bytes))
                texts   = [page.extract_text() or "" for page in reader.pages[:20]]
                content = "\n\n---PAGE---\n\n".join(texts)[:8000]
                summary = f"PDF: {len(reader.pages)} pages"
                metadata = {"pages": len(reader.pages)}
            except ImportError:
                summary = "PDF file — install PyPDF2 to mine data"
            except Exception as ex:
                summary = f"PDF parse error: {ex}"

        else:
            summary = f"Binary file ({ext}), {len(raw_bytes)} bytes"
    except Exception as e:
        summary = f"Mining error: {e}"

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO dropped_files (id,filename,content,file_type,summary,metadata) VALUES (?,?,?,?,?,?)",
            (fid, filename, content, ext, summary, json.dumps(metadata))
        )
        await db.commit()

    # Return mined data for immediate use
    return {
        "id":       fid,
        "filename": filename,
        "summary":  summary,
        "metadata": metadata,
        "preview":  content[:2000] if content else "",
        "size":     len(raw_bytes),
    }


@router.get("/api/files/dropped")
async def get_dropped_files(limit: int = 50):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT id,filename,file_type,summary,metadata,created_at FROM dropped_files ORDER BY created_at DESC LIMIT ?",
            (limit,)
        ) as cur:
            rows = [dict(r) for r in await cur.fetchall()]
    for r in rows:
        try: r["metadata"] = json.loads(r["metadata"])
        except: pass
    return rows


@router.get("/api/files/dropped/{fid}")
async def get_dropped_file(fid: str):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM dropped_files WHERE id=?", (fid,)) as cur:
            row = await cur.fetchone()
    if not row:
        return {"error": "Not found"}
    r = dict(row)
    try: r["metadata"] = json.loads(r["metadata"])
    except: pass
    return r
